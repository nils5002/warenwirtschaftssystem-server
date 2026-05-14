from __future__ import annotations

from io import BytesIO
from uuid import uuid4

from fastapi.testclient import TestClient

from app.main import app
from .auth_helpers import auth_headers


def _headers() -> dict[str, str]:
    return auth_headers(TestClient(app), "Admin")


def _build_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Event Laptops"])
    sheet.append(
        [
            "Name",
            "Modell",
            "Seriennummer",
            "IP-Adresse",
            "Mac-Adresse LAN",
            "Mac-Adresse WLAN",
            "Kategorie",
            "Notizen",
        ]
    )
    suffix = uuid4().hex[:6].upper()
    sheet.append(
        [
            f"CX-EVENT-{suffix}",
            "Lenovo T14",
            f"PF-{suffix}",
            "192.168.10.141",
            "90-2E-16-19-CF-24",
            "F4-4E-E3-96-DC-E6",
            "Notebook",
            "Import-Test",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_simple_ipad_workbook_bytes() -> tuple[bytes, set[str]]:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    suffix = uuid4().hex[:6].upper()
    serial_one = f"XK20{suffix}A"
    serial_two = f"TWCF{suffix}B"
    sheet.append(["iPad", "Seriennummer", "Mac-Adresse"])
    sheet.append([1, serial_one, "60:DD:70:AA:23:D5"])
    sheet.append([2, serial_two, "10:9F:41:3B:8F:C8"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), {serial_one, serial_two}


def _build_unknown_category_header_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["MysteryDevice", "Seriennummer", "Mac-Adresse"])
    sheet.append([1, "SN-UNKNOWN-01", "AA:BB:CC:DD:EE:01"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_event_handheld_like_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Nummer", "Seriennummer", "Sim Karten Nummer", "Mac-Adresse", "Status"])
    sheet.append([1, "357585120047645", "", "00:26:E8:0A:11:22", "OK"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_event_qr_like_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Modell", "Sprache", "Status"])
    sheet.append(["", "Zebra DS2278", "Deutsch", "OK"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_laserdrucker_like_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Modell", "IP-Adresse", "Mac-Adresse WLAN", "Mac-Adresse LAN", "Netzteil", "Status"])
    sheet.append(["", "HP LaserJet 4200", "192.168.50.20", "", "90-2E-16-19-CF-24", "Kaltgeräte", "OK"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_no_serial_mac_only_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    # Use a unique MAC per test run to avoid IntegrityError on repeated runs
    unique_hex = uuid4().hex[:12].upper()
    mac = ":".join(unique_hex[i : i + 2] for i in range(0, 12, 2))
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Handheld", "Mac-Adresse"])
    sheet.append([7, mac])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_upload_preview_and_confirm_flow() -> None:
    client = TestClient(app)
    payload = _build_workbook_bytes()

    preview = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={"file": ("hardware_import.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    preview_json = preview.json()
    assert preview_json["rows_valid"] >= 1
    assert preview_json["recognized_columns"]

    confirm = client.post(
        "/api/wms/import/confirm",
        headers=_headers(),
        json={"preview_id": preview_json["preview_id"]},
    )
    assert confirm.status_code == 200
    confirm_json = confirm.json()
    assert confirm_json["imported_count"] + confirm_json["updated_count"] >= 1


def test_upload_rejects_non_excel() -> None:
    client = TestClient(app)
    response = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={"file": ("bad.csv", b"name;serial", "text/csv")},
    )
    assert response.status_code == 400


def test_upload_preview_accepts_simple_ipad_list_and_generates_names() -> None:
    client = TestClient(app)
    payload, expected_serials = _build_simple_ipad_workbook_bytes()

    preview = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={"file": ("Genolive iPads.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    data = preview.json()
    assert data["rows_valid"] == 2
    assert data["unresolved_category_rows"] == 0
    assert data["column_mapping"].get("category_source") == "ipad"
    assert data["column_mapping"].get("serial_number") == "seriennummer"
    assert data["column_mapping"].get("mac_generic") == "mac adresse"
    assert data["auto_generated_names"] == 2
    assert any("automatisch aus Kategorie + Nummer" in item for item in data["warnings"])

    confirm = client.post(
        "/api/wms/import/confirm",
        headers=_headers(),
        json={"preview_id": data["preview_id"]},
    )
    assert confirm.status_code == 200
    result = confirm.json()
    assert result["imported_count"] + result["updated_count"] >= 2

    assets = client.get("/api/wms/assets", headers=_headers())
    assert assets.status_code == 200
    payload_assets = assets.json()
    imported = [item for item in payload_assets if item.get("serialNumber") in expected_serials]
    assert len(imported) == 2
    assert {item["name"] for item in imported} == {"iPad 1", "iPad 2"}
    assert all(item["category"] == "iPad" for item in imported)
    assert all(item["status"] in {"Verfuegbar", "Verfügbar"} for item in imported)
    observed_wlan = {item.get("macAddressWlan") or item.get("macWlan") for item in imported}
    assert observed_wlan == {"60:DD:70:AA:23:D5", "10:9F:41:3B:8F:C8"}


def test_upload_preview_accepts_user_defined_db_category() -> None:
    """Selbstdefinierte Kategorien (aus dem Kategorien-Modul) muessen vom
    Excel-Importer akzeptiert werden — nicht nur die hartcodierten Standards.

    Frueher haben Excel-Zeilen mit einer nicht-Standard-Kategorie pauschal
    den Validator-Fehler "Kategorie-Zuordnung erforderlich" produziert,
    auch wenn der Operator die Kategorie zuvor sauber im UI angelegt hatte
    (siehe DYMO-Importfall). Dieser Test sichert das Verhalten ab.
    """
    from openpyxl import Workbook

    client = TestClient(app)
    suffix = uuid4().hex[:6].upper()
    custom_category = f"CustomCat-{suffix}"

    # Kategorie zuerst ueber das offizielle API anlegen, damit der Datenpfad
    # exakt dem Operator-Flow entspricht.
    created = client.post(
        "/api/wms/categories",
        headers=_headers(),
        json={"name": custom_category, "isActive": True},
    )
    assert created.status_code in {200, 201}, created.text

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([custom_category])  # Titelzeile
    sheet.append(["Name", "Modell", "Kategorie"])
    sheet.append([f"CX-{suffix}-1", "Custom Device A", custom_category])
    sheet.append([f"CX-{suffix}-2", "Custom Device B", custom_category])
    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()

    preview = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={
            "file": (
                f"{custom_category}.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview.status_code == 200, preview.text
    data = preview.json()
    assert data["rows_total"] == 2
    assert data["rows_valid"] == 2
    assert data["unresolved_category_rows"] == 0
    assert data["errors"] == []

    confirm = client.post(
        "/api/wms/import/confirm",
        headers=_headers(),
        json={"preview_id": data["preview_id"]},
    )
    assert confirm.status_code == 200, confirm.text
    result = confirm.json()
    assert result["imported_count"] + result["updated_count"] == 2

    assets = client.get("/api/wms/assets", headers=_headers())
    assert assets.status_code == 200
    imported = [
        item for item in assets.json()
        if item.get("name", "").startswith(f"CX-{suffix}-")
    ]
    assert len(imported) == 2
    assert all(item["category"] == custom_category for item in imported)


def test_upload_preview_marks_unknown_category_header_as_unresolved() -> None:
    client = TestClient(app)
    payload = _build_unknown_category_header_workbook_bytes()

    preview = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={"file": ("unknown_devices.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    data = preview.json()
    assert data["rows_valid"] == 0
    assert data["unresolved_category_rows"] == 1


def test_upload_preview_maps_event_handheld_variant() -> None:
    client = TestClient(app)
    payload = _build_event_handheld_like_workbook_bytes()
    preview = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={"file": ("event_handhelden.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    data = preview.json()
    assert data["inferred_category"] == "Handheld"
    assert data["rows_valid"] == 1


def test_upload_preview_maps_event_qr_variant_to_qr_scanner() -> None:
    client = TestClient(app)
    payload = _build_event_qr_like_workbook_bytes()
    preview = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={"file": ("event_qrcodescan.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    data = preview.json()
    assert data["inferred_category"] == "QR-Code-Scanner"
    assert data["unresolved_category_rows"] == 0


def test_upload_preview_maps_laserdrucker_variant_with_ip_and_lan_mac() -> None:
    client = TestClient(app)
    payload = _build_laserdrucker_like_workbook_bytes()
    preview = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={
            "file": ("Genolive Laserdrucker.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        },
    )
    assert preview.status_code == 200
    data = preview.json()
    assert data["inferred_category"] == "Drucker"
    confirm = client.post(
        "/api/wms/import/confirm",
        headers=_headers(),
        json={"preview_id": data["preview_id"]},
    )
    assert confirm.status_code == 200


def test_upload_without_serial_uses_deterministic_auto_serial_and_no_duplicate_on_reimport() -> None:
    client = TestClient(app)
    payload = _build_no_serial_mac_only_workbook_bytes()

    preview_one = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={"file": ("handheld_no_serial.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview_one.status_code == 200
    data_one = preview_one.json()
    assert data_one["auto_generated_serials"] >= 1
    confirm_one = client.post(
        "/api/wms/import/confirm",
        headers=_headers(),
        json={"preview_id": data_one["preview_id"]},
    )
    assert confirm_one.status_code == 200
    first = confirm_one.json()
    assert first["imported_count"] + first["updated_count"] >= 1

    preview_two = client.post(
        "/api/wms/import/preview",
        headers=_headers(),
        files={"file": ("handheld_no_serial.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview_two.status_code == 200
    data_two = preview_two.json()
    confirm_two = client.post(
        "/api/wms/import/confirm",
        headers=_headers(),
        json={"preview_id": data_two["preview_id"]},
    )
    assert confirm_two.status_code == 200
    second = confirm_two.json()
    assert second["imported_count"] == 0
