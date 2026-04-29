from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from sqlalchemy import select

from app.database.models import AssetRecord
from app.database.session import SessionLocal
from app.services.excel_import_service import _is_effectively_empty_row, _prepare_row_data
from app.services.hardware_import.importer import upsert_asset_by_serial
from app.services.hardware_import.mapper import map_excel_row_to_asset
from app.services.hardware_import.parser import parse_excel_file
from app.services.hardware_import.types import ParsedExcelRow
from app.services.hardware_import.validator import validate_mapped_payload


def test_parser_skips_title_row_and_detects_flexible_headers(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Event Laptops"
    sheet.append(["Event Laptops"])
    sheet.append(["Name", "Modell", "S/N", "IP-Adresse", "Mac-Adresse LAN", "WLAN-MAC", "Status"])
    sheet.append(
        [
            "CX-EVENT-01",
            "Lenovo T14",
            "PF-2YA4ZY",
            "192.168.10.141",
            "90-2E-16-19-CF-24",
            "F4-4E-E3-96-DC-E6",
            "OK",
        ]
    )
    file_path = tmp_path / "event_laptops.xlsx"
    workbook.save(file_path)

    parsed = parse_excel_file(file_path)

    assert parsed.rows
    assert parsed.rows[0].row_number == 3
    assert parsed.recognized_columns == [
        "name",
        "model",
        "serial_number",
        "ip_address",
        "mac_lan",
        "mac_wlan",
        "status",
    ]


def test_parser_detects_category_from_header_for_simple_ipad_list(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Genolive iPads"
    sheet.append(["iPad", "Seriennummer", "Mac-Adresse"])
    sheet.append([1, "XK20R2QP7C", "60:DD:70:AA:23:D5"])
    file_path = tmp_path / "genolive_ipads.xlsx"
    workbook.save(file_path)

    parsed = parse_excel_file(file_path)

    assert parsed.rows
    assert "category_source" in parsed.recognized_columns
    assert parsed.column_mapping.get("category_source") == "ipad"


def test_mapper_normalizes_notebooks_to_laptop_and_keeps_excel_status_as_legacy_note() -> None:
    row = ParsedExcelRow(
        file_name="Event_Laptops.xlsx",
        sheet_name="Sheet1",
        row_number=2,
        data={
            "name": "CX-EVENT-01",
            "model": "Lenovo T14",
            "serial_number": "PF-2YA4ZY",
            "ip_address": "192.168.10.141",
            "mac_lan": "90-2E-16-19-CF-24",
            "mac_wlan": "F4-4E-E3-96-DC-E6",
            "status": "OK",
        },
    )

    mapped = map_excel_row_to_asset(row)

    assert mapped.payload["category"] == "Laptop"
    assert mapped.payload["status"] == "Verfuegbar"
    assert "Import-Status (Legacy): OK" in mapped.payload["notes"]


def test_mapper_generates_name_and_maps_generic_mac_to_wlan_for_ipad() -> None:
    row = ParsedExcelRow(
        file_name="Genolive iPads.xlsx",
        sheet_name="Sheet1",
        row_number=2,
        data={
            "inventory_number": 1,
            "category_source": "iPad",
            "serial_number": "XK20R2QP7C",
            "mac_generic": "60:DD:70:AA:23:D5",
        },
    )

    mapped = map_excel_row_to_asset(row)

    assert mapped.payload["category"] == "iPad"
    assert mapped.payload["name"] == "iPad 1"
    assert mapped.payload["mac_wlan"] == "60:DD:70:AA:23:D5"
    assert mapped.payload["mac_lan"] is None
    assert mapped.payload["status"] == "Verfuegbar"
    assert mapped.auto_generated_name is True


def test_mapper_maps_event_handheld_variant() -> None:
    row = ParsedExcelRow(
        file_name="event_handhelden.xlsx",
        sheet_name="Tabelle1",
        row_number=2,
        data={
            "inventory_number": 1,
            "serial_number": "357585120047645",
            "sim_number": "",
            "mac_generic": "00:26:E8:0A:11:22",
            "status": "OK",
        },
    )
    mapped = map_excel_row_to_asset(row)
    assert mapped.payload["category"] == "Handheld"
    assert mapped.payload["name"] == "Handheld 1"
    assert mapped.payload["mac_wlan"] == "00:26:E8:0A:11:22"
    assert mapped.payload["status"] == "Verfuegbar"


def test_mapper_maps_qr_codescan_variant_not_as_handheld() -> None:
    row = ParsedExcelRow(
        file_name="event_qrcodescan.xlsx",
        sheet_name="Tabelle1",
        row_number=2,
        data={
            "name": "",
            "inventory_number": 5,
            "model": "Zebra DS2278",
            "status": "OK",
        },
    )
    mapped = map_excel_row_to_asset(row)
    assert mapped.payload["category"] == "QR-Code-Scanner"
    assert mapped.payload["name"] == "QR-Code-Scanner 5"


def test_mapper_maps_laserdrucker_variant_prefers_lan_mac() -> None:
    row = ParsedExcelRow(
        file_name="Genolive Laserdrucker.xlsx",
        sheet_name="Tabelle1",
        row_number=2,
        data={
            "name": "",
            "model": "HP LaserJet 4200",
            "ip_address": "192.168.1.42",
            "mac_lan": "",
            "mac_wlan": "",
            "mac_generic": "90-2E-16-19-CF-24",
            "status": "OK",
        },
    )
    mapped = map_excel_row_to_asset(row)
    assert mapped.payload["category"] == "Drucker"
    assert mapped.payload["ip_address"] == "192.168.1.42"
    assert mapped.payload["mac_lan"] == "90:2E:16:19:CF:24"
    assert mapped.payload["mac_wlan"] is None
    assert mapped.payload["device_model"] == "HP LaserJet 4200"


def test_mapper_generates_deterministic_auto_serial_from_mac_when_missing_serial() -> None:
    row = ParsedExcelRow(
        file_name="event_handhelden.xlsx",
        sheet_name="Tabelle1",
        row_number=7,
        data={
            "inventory_number": 7,
            "serial_number": "",
            "mac_generic": "10:9F:41:3B:8F:C8",
        },
    )
    mapped_one = map_excel_row_to_asset(row)
    mapped_two = map_excel_row_to_asset(row)
    assert mapped_one.serial_number.startswith("AUTO-HANDHELD-")
    assert mapped_one.serial_number == mapped_two.serial_number
    assert mapped_one.auto_generated_serial is True


def test_importer_falls_back_to_name_and_category_for_auto_serial_duplicates() -> None:
    suffix = uuid4().hex[:8]
    seed_external_id = f"asset-import-{suffix}"
    unique_name = f"CX-EVENT-{suffix}"
    existing = AssetRecord(
        external_id=seed_external_id,
        name=unique_name,
        category="Laptop",
        location="Eventlager",
        status="Verfuegbar",
        assigned_to="-",
        next_return="-",
        tag_number=f"IMP-{suffix}",
        serial_number=f"PF-{suffix}",
        device_model="Lenovo T14",
        ip_address=None,
        mac_lan=None,
        mac_wlan=None,
        qr_code=f"WMS|{seed_external_id}|IMP-{suffix}",
        maintenance_state="Importiert",
        notes="Seed",
        last_checkout="-",
        next_reservation="-",
        source_file="seed.xlsx",
    )

    with SessionLocal() as db:
        db.add(existing)
        db.commit()

        payload = {
            "external_id": f"asset-auto-{suffix}",
            "name": unique_name,
            "category": "Laptop",
            "location": "Eventlager",
            "status": "Verfuegbar",
            "assigned_to": "-",
            "next_return": "-",
            "tag_number": f"IMP-AUTO-{suffix}",
            "serial_number": f"AUTO-{suffix.upper()}",
            "device_model": "Lenovo T14 Gen 5",
            "ip_address": None,
            "mac_lan": None,
            "mac_wlan": None,
            "qr_code": f"WMS|asset-auto-{suffix}|IMP-AUTO-{suffix}",
            "maintenance_state": "Importiert",
            "notes": "Import",
            "last_checkout": "-",
            "next_reservation": "-",
            "source_file": "import.xlsx",
        }

        action = upsert_asset_by_serial(db, payload, dry_run=False)
        db.commit()

        assert action.action == "updated"
        refreshed = db.scalar(select(AssetRecord).where(AssetRecord.external_id == seed_external_id))
        assert refreshed is not None
        assert refreshed.device_model == "Lenovo T14 Gen 5"


# ---------------------------------------------------------------------------
# Parser integration tests for the three real-world file variants
# ---------------------------------------------------------------------------

def _run_full_pipeline(parsed_file) -> tuple[int, int, int, list[str]]:
    """Map + validate every row. Returns (valid, auto_names, auto_serials, errors)."""
    valid, auto_names, auto_serials = 0, 0, 0
    errors: list[str] = []
    seen: set[str] = set()
    for row in parsed_file.rows:
        if _is_effectively_empty_row(row.data):
            continue
        row_data = _prepare_row_data(row.data, row.file_name, row.row_number)
        mapped = map_excel_row_to_asset(
            ParsedExcelRow(
                file_name=parsed_file.file_name,
                sheet_name=parsed_file.sheet_name,
                row_number=row.row_number,
                data=row_data,
            )
        )
        errs = validate_mapped_payload(mapped.payload)
        if errs:
            errors.extend(errs)
            continue
        if mapped.serial_number in seen:
            errors.append(f"Duplikat: {mapped.serial_number}")
            continue
        seen.add(mapped.serial_number)
        valid += 1
        if mapped.auto_generated_name:
            auto_names += 1
        if mapped.auto_generated_serial:
            auto_serials += 1
    return valid, auto_names, auto_serials, errors


def test_parser_handheld_structure_no_name_column(tmp_path: Path) -> None:
    """Handheld file: Nummer/Seriennummer/Sim/Mac-Adresse/Status – no Name column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelle1"
    ws.append(["Nummer", "Seriennummer", "Sim Karten Nummer", "Mac-Adresse", "Status"])
    ws.append([1, "357585120047645", None, "b4:29:3d:99:63:e6", "OK"])
    ws.append([2, "357585120047646", None, "b4:29:3d:99:63:e7", "OK"])
    ws.append([3, "357585120047647", None, "b4:29:3d:99:63:e8", "OK"])
    path = tmp_path / "event_handhelden.xlsx"
    wb.save(path)

    parsed = parse_excel_file(path)

    assert parsed.inferred_category == "Handheld"
    assert parsed.inferred_category_source == "file_name"
    assert "serial_number" in parsed.recognized_columns
    assert "mac_generic" in parsed.recognized_columns
    assert "inventory_number" in parsed.recognized_columns
    assert parsed.missing_required_columns == []

    valid, auto_names, auto_serials, errors = _run_full_pipeline(parsed)
    assert errors == [], f"Unexpected errors: {errors}"
    assert valid == 3
    assert auto_serials == 0  # real serials provided
    assert auto_names == 3    # no Name column → auto-generated


def test_parser_handheld_mac_routes_to_wlan(tmp_path: Path) -> None:
    """Handheld mac_generic → mac_wlan (Handheld prefers WLAN)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelle1"
    ws.append(["Nummer", "Seriennummer", "Mac-Adresse", "Status"])
    ws.append([1, "SN-HAND-001", "b4:29:3d:99:63:e6", "OK"])
    path = tmp_path / "event_handhelden.xlsx"
    wb.save(path)

    parsed = parse_excel_file(path)
    row = parsed.rows[0]
    row_data = _prepare_row_data(row.data, row.file_name, row.row_number)
    mapped = map_excel_row_to_asset(
        ParsedExcelRow(file_name=parsed.file_name, sheet_name=parsed.sheet_name,
                       row_number=row.row_number, data=row_data)
    )

    assert mapped.payload["mac_wlan"] == "B4:29:3D:99:63:E6"
    assert mapped.payload["mac_lan"] is None
    assert mapped.payload["category"] == "Handheld"


def test_parser_qrscanner_two_sections_duplicate_header(tmp_path: Path) -> None:
    """QR scanner file: two sections separated by empty row + duplicate header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelle1"
    # Section 1
    ws.append(["Name", "Modell", "Sprache", "Status"])
    ws.append(["CX-QRUSB-01", "Albasca MK-7000-2D", "Deutsch", "OK"])
    ws.append(["CX-QRUSB-02", "Albasca MK-7000-2D", "Deutsch", "OK"])
    # Empty separator
    ws.append([None, None, None, None])
    # Duplicate header
    ws.append(["Name", "Modell", "Sprache", "Status"])
    # Section 2 – no names
    ws.append([None, "Sumeber Barcode-Scanner 2D", "Deutsch", "OK"])
    ws.append([None, "Sumeber Barcode-Scanner 2D", "Deutsch", "OK"])
    ws.append([None, "Sumeber Barcode-Scanner 2D", "Deutsch", "OK"])
    path = tmp_path / "event_qrcodescan.xlsx"
    wb.save(path)

    parsed = parse_excel_file(path)

    assert parsed.inferred_category == "QR-Code-Scanner"
    # Duplicate header row must be skipped → only 5 data rows (2 section1 + 3 section2)
    assert len(parsed.rows) == 5

    valid, auto_names, auto_serials, errors = _run_full_pipeline(parsed)
    assert errors == [], f"Unexpected errors: {errors}"
    assert valid == 5
    assert auto_serials == 5   # no serial column
    assert auto_names == 3     # section 2 has no names


def test_parser_qrscanner_not_classified_as_handheld(tmp_path: Path) -> None:
    """QR scanner file must not be mis-classified as Handheld."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelle1"
    ws.append(["Name", "Modell", "Status"])
    ws.append(["CX-QR-01", "Zebra DS2278", "OK"])
    path = tmp_path / "event_qrcodescan.xlsx"
    wb.save(path)

    parsed = parse_excel_file(path)
    assert parsed.inferred_category == "QR-Code-Scanner"
    assert parsed.inferred_category != "Handheld"


def test_parser_laserdrucker_structure_no_serial_col(tmp_path: Path) -> None:
    """Laserdrucker file: Name/IP/MAC LAN/MAC WLAN/Netzteil/Status, no serial, duplicate header mid-sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelle1"
    ws.append(["Name", "Modell", "IP-Adresse", "Mac-Adresse LAN", "Mac-Adresse WLAN", "Netzteil", "Status"])
    ws.append(["CX-LASER-01", "HP LaserJet Pro 400", "192.168.10.211", "d0-bf-9c-bf-46-81", None, "Kaltgeraete", "OK"])
    ws.append(["CX-LASER-02", "HP LaserJet Pro 400", "192.168.10.212", "d0-bf-9c-bf-46-82", None, "Kaltgeraete", "OK"])
    ws.append(["CX-LASER-03", "Brother HL-L5100DN", "192.168.10.220", "c4-34-6b-11-22-33", None, "Kaltgeraete", "OK"])
    # Duplicate header (simulates Kyocera section)
    ws.append(["Name", "Modell", "IP-Adresse", "Mac-Adresse LAN", "Mac-Adresse WLAN", "Netzteil", "Status"])
    ws.append(["CX-LASER-04", "Kyocera PA2100cx", "192.168.10.230", None, None, "Kaltgeraete", "OK"])
    path = tmp_path / "Genolive Laserdrucker.xlsx"
    wb.save(path)

    parsed = parse_excel_file(path)

    assert parsed.inferred_category == "Drucker"
    assert parsed.inferred_category_source == "file_name"
    assert "ip_address" in parsed.recognized_columns
    assert "mac_lan" in parsed.recognized_columns
    assert "mac_wlan" in parsed.recognized_columns
    # Duplicate header must be skipped → 4 data rows
    assert len(parsed.rows) == 4

    valid, auto_names, auto_serials, errors = _run_full_pipeline(parsed)
    assert errors == [], f"Unexpected errors: {errors}"
    assert valid == 4
    assert auto_serials == 4  # no serial column


def test_parser_laserdrucker_mac_lan_preferred_for_printer(tmp_path: Path) -> None:
    """For Drucker, mac_generic goes to mac_lan (not mac_wlan)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelle1"
    ws.append(["Name", "Modell", "IP-Adresse", "Mac-Adresse", "Status"])
    ws.append(["CX-LASER-01", "HP LaserJet", "192.168.1.10", "d0-bf-9c-bf-46-81", "OK"])
    path = tmp_path / "Genolive Laserdrucker.xlsx"
    wb.save(path)

    parsed = parse_excel_file(path)
    row = parsed.rows[0]
    row_data = _prepare_row_data(row.data, row.file_name, row.row_number)
    mapped = map_excel_row_to_asset(
        ParsedExcelRow(file_name=parsed.file_name, sheet_name=parsed.sheet_name,
                       row_number=row.row_number, data=row_data)
    )

    assert mapped.payload["mac_lan"] == "D0:BF:9C:BF:46:81"
    assert mapped.payload["mac_wlan"] is None
    assert mapped.payload["ip_address"] == "192.168.1.10"
    assert mapped.payload["category"] == "Drucker"


def test_reimport_no_serial_column_is_idempotent(tmp_path: Path) -> None:
    """Re-importing a file without serial column must produce identical AUTO-serials (stable hashes)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelle1"
    ws.append(["Name", "Modell", "IP-Adresse", "Status"])
    ws.append(["CX-LASER-01", "HP LaserJet", "192.168.1.10", "OK"])
    path = tmp_path / "Genolive Laserdrucker.xlsx"
    wb.save(path)

    parsed1 = parse_excel_file(path)
    parsed2 = parse_excel_file(path)

    def get_serials(parsed):
        serials = []
        for row in parsed.rows:
            row_data = _prepare_row_data(row.data, row.file_name, row.row_number)
            mapped = map_excel_row_to_asset(
                ParsedExcelRow(file_name=parsed.file_name, sheet_name=parsed.sheet_name,
                               row_number=row.row_number, data=row_data)
            )
            serials.append(mapped.serial_number)
        return serials

    serials1 = get_serials(parsed1)
    serials2 = get_serials(parsed2)

    assert serials1 == serials2
    assert all(s.startswith("AUTO-") for s in serials1)
