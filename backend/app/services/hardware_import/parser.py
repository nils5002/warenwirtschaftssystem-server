from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .categorizer import infer_category_with_source, normalize_category_label
from .types import CANONICAL_COLUMNS, ParsedExcelFile, ParsedExcelRow, REQUIRED_COLUMNS

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "name": (
        "name",
        "geraetename",
        "gerätename",
        "gerät",
        "geraet",
        "bezeichnung",
    ),
    "category": ("kategorie", "category", "geraetekategorie", "gerätekategorie", "asset category"),
    "inventory_number": ("nummer", "id", "inventarnummer", "inventar nummer", "inventory number", "asset number"),
    "model": ("modell", "model", "gerätetyp", "geraetetyp", "typ", "type", "produkt", "produktname"),
    "description": ("beschreibung", "description", "kommentar", "details", "bemerkung", "notes"),
    "serial_number": (
        "seriennummer",
        "serial",
        "serial number",
        "serialnumber",
        "s/n",
        "sn",
        "serien nr",
        "serien-nr",
        "seriennr",
        "serien nr.",
    ),
    "sim_number": ("sim karten nummer", "sim-kartennummer", "sim kartennummer", "sim-karte", "sim"),
    "phone_number": ("rufnummer", "telefonnummer", "telefon", "phone", "phone number"),
    "power_supply": ("netzteil", "ladegerät", "ladegeraet", "zubehör", "zubehoer", "accessory", "power supply"),
    "language": ("sprache", "language"),
    "ip_address": ("ip", "ipadresse", "ip-adresse", "ip adresse", "ipv4", "netzwerkadresse"),
    "mac_lan": (
        "mac lan",
        "mac_lan",
        "maclan",
        "mac adresse lan",
        "mac-adresse lan",
        "mac-address lan",
        "lan mac",
        "ethernet mac",
        "rj45 mac",
    ),
    "mac_wlan": (
        "mac wlan",
        "mac_wlan",
        "macwlan",
        "mac adresse wlan",
        "mac-adresse wlan",
        "mac-address wlan",
        "wlan mac",
        "wifi mac",
        "wlan-mac",
        "wireless mac",
    ),
    "mac_generic": ("mac", "macadresse", "mac adresse", "mac-adresse", "mac address", "hardwareadresse"),
    "status": ("status", "zustand"),
}


def list_excel_files(import_dir: Path) -> tuple[list[Path], list[str]]:
    files: list[Path] = []
    skipped: list[str] = []
    for path in sorted(import_dir.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file():
            continue
        if path.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(path)
        else:
            skipped.append(path.name)
    return files, skipped


def parse_excel_file(path: Path, *, original_file_name: str | None = None) -> ParsedExcelFile:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        display_name = original_file_name or path.name
        parsed = ParsedExcelFile(path=path, file_name=display_name, sheet_name=sheet.title)

        row_iter = sheet.iter_rows(values_only=True)
        buffered_rows = _take_rows(row_iter, limit=40)
        header_index, col_idx_by_canonical, column_mapping = find_header_row(buffered_rows)
        if header_index is None:
            parsed.missing_required_columns = list(REQUIRED_COLUMNS)
            return parsed

        parsed.missing_required_columns = [col for col in REQUIRED_COLUMNS if col not in col_idx_by_canonical]
        parsed.missing_optional_columns = [
            col for col in CANONICAL_COLUMNS if col not in col_idx_by_canonical and col not in REQUIRED_COLUMNS
        ]
        parsed.recognized_columns = sorted(col_idx_by_canonical, key=lambda key: col_idx_by_canonical[key])
        parsed.column_mapping = column_mapping
        parsed.title_hint = extract_title_hint(buffered_rows, header_index)

        file_stem = Path(display_name).stem
        category, source, source_label = infer_category_with_source(
            explicit_category=None,
            header_category=column_mapping.get("category_source"),
            sheet_name=sheet.title,
            file_name=file_stem,
            title_hint=parsed.title_hint,
            name=file_stem,
            model=None,
            description=None,
        )
        parsed.inferred_category = category
        parsed.inferred_category_source = source
        parsed.inferred_category_source_label = source_label

        header_row_pos = next(
            (idx for idx, (row_number, _) in enumerate(buffered_rows) if row_number == header_index),
            None,
        )
        if header_row_pos is None:
            parsed.missing_required_columns = list(REQUIRED_COLUMNS)
            return parsed

        for row_number, raw_row in buffered_rows[header_row_pos + 1 :]:
            _append_parsed_row(parsed, sheet.title, row_number, raw_row, col_idx_by_canonical)

        row_number = (buffered_rows[-1][0] + 1) if buffered_rows else 1
        for raw_row in row_iter:
            if is_header_repeat(raw_row, col_idx_by_canonical):
                row_number += 1
                continue
            _append_parsed_row(parsed, sheet.title, row_number, raw_row, col_idx_by_canonical)
            row_number += 1
        return parsed
    finally:
        workbook.close()


def _append_parsed_row(
    parsed: ParsedExcelFile,
    sheet_name: str,
    row_number: int,
    raw_row: tuple[object, ...],
    col_idx_by_canonical: dict[str, int],
) -> None:
    if is_header_repeat(raw_row, col_idx_by_canonical):
        return
    data: dict[str, object] = {}
    for canonical, idx in col_idx_by_canonical.items():
        if canonical == "category_source":
            data[canonical] = parsed.column_mapping.get("category_source")
            continue
        value = raw_row[idx] if idx < len(raw_row) else None
        data[canonical] = value
    data["_sheet_name"] = sheet_name
    if parsed.title_hint:
        data["_title_hint"] = parsed.title_hint
    if any(value not in (None, "") for key, value in data.items() if not str(key).startswith("_")):
        parsed.rows.append(
            ParsedExcelRow(
                file_name=parsed.file_name,
                sheet_name=sheet_name,
                row_number=row_number,
                data=data,
            )
        )


def extract_title_hint(buffered_rows: list[tuple[int, tuple[object, ...]]], header_index: int) -> str | None:
    for row_number, row_values in buffered_rows:
        if row_number >= header_index:
            break
        non_empty = [str(value).strip() for value in row_values if str(value or "").strip()]
        if len(non_empty) == 1:
            title = non_empty[0]
            if normalize_category_label(title):
                return title
    return None


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("_", " ").replace("-", " ").replace("/", " ")
    text = re.sub(r"[^a-z0-9äöüß ]+", " ", text)
    return " ".join(text.split())


NORMALIZED_ALIAS_MAP: dict[str, set[str]] = {
    canonical: {normalize_header(alias) for alias in aliases}
    for canonical, aliases in COLUMN_ALIASES.items()
}


def resolve_columns(header: list[str]) -> tuple[dict[str, int], dict[str, str]]:
    mapping: dict[str, int] = {}
    column_mapping: dict[str, str] = {}

    for canonical, alias_set in NORMALIZED_ALIAS_MAP.items():
        for idx, header_name in enumerate(header):
            if header_name in alias_set:
                mapping[canonical] = idx
                column_mapping[canonical] = header_name
                break

    if "category" not in mapping and "category_source" not in mapping:
        for idx, header_name in enumerate(header):
            if not header_name:
                continue
            normalized_category = normalize_category_label(header_name)
            if normalized_category:
                mapping["category_source"] = idx
                column_mapping["category_source"] = header_name
                if "inventory_number" not in mapping:
                    mapping["inventory_number"] = idx
                    column_mapping["inventory_number"] = header_name
                break

    return mapping, column_mapping


def find_header_row(
    buffered_rows: list[tuple[int, tuple[object, ...]]]
) -> tuple[int | None, dict[str, int], dict[str, str]]:
    best_row_index: int | None = None
    best_mapping: dict[str, int] = {}
    best_column_mapping: dict[str, str] = {}
    best_score = -1
    best_required = -1

    for row_index, row_values in buffered_rows:
        header = [normalize_header(value) for value in row_values]
        if not any(header):
            continue
        mapping, column_mapping = resolve_columns(header)
        score = len(mapping)
        required_matches = sum(1 for col in REQUIRED_COLUMNS if col in mapping)
        if score > best_score or (score == best_score and required_matches > best_required):
            best_score = score
            best_required = required_matches
            best_row_index = row_index
            best_mapping = mapping
            best_column_mapping = column_mapping

    if best_row_index is None or best_score <= 0:
        return None, {}, {}
    identity_columns = {"serial_number", "name", "model", "ip_address", "mac_lan", "mac_wlan", "mac_generic"}
    if not (identity_columns & set(best_mapping)):
        return None, {}, {}
    return best_row_index, best_mapping, best_column_mapping


def is_header_repeat(raw_row: tuple[object, ...], col_idx_by_canonical: dict[str, int]) -> bool:
    checks = 0
    matches = 0
    for canonical, idx in col_idx_by_canonical.items():
        if canonical == "category_source":
            continue
        value = normalize_header(raw_row[idx] if idx < len(raw_row) else "")
        if not value:
            continue
        checks += 1
        if value in NORMALIZED_ALIAS_MAP.get(canonical, set()):
            matches += 1
    return checks > 0 and matches == checks


def _take_rows(
    row_iter: Iterable[tuple[object, ...]],
    *,
    limit: int,
) -> list[tuple[int, tuple[object, ...]]]:
    rows: list[tuple[int, tuple[object, ...]]] = []
    for row_number, row in enumerate(row_iter, start=1):
        rows.append((row_number, row))
        if row_number >= limit:
            break
    return rows
